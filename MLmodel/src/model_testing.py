from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
OUTPUT_DIR = REPO_ROOT / "MLmodel" / "MLfiles"
QUALITY_BINS = [0.05, 0.20, 0.50, 0.80]
QUALITY_LABELS = ["Excellent", "Good", "Fair", "Poor", "Critical"]
PRIORITY_MAP = {
    "Critical": 1,
    "Poor": 2,
    "Fair": 3,
    "Good": 4,
    "Excellent": 5,
}
PRIORITY_LABELS = ["Poor", "Critical"]
MISMATCH_Q = 0.10


def _apply_data_sheet_formatting(worksheet, data_columns: list[str], max_row: int) -> None:
    if max_row < 2:
        return

    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font

    worksheet.freeze_panes = "E2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(data_columns))}{max_row}"

    color_scale = ColorScaleRule(
        start_type="min",
        start_color="63BE7B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="F8696B",
    )
    gradient_columns = [
        "pys_kiiht_raw",
        "siv_kiiht_raw",
        "nyo_kiiht_raw",
        "anomaly_score",
    ]
    col_index = {name: idx + 1 for idx, name in enumerate(data_columns)}
    for column_name in gradient_columns:
        if column_name not in col_index:
            continue
        col_letter = get_column_letter(col_index[column_name])
        worksheet.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{max_row}",
            color_scale,
        )

    if "mismatch_flag" in col_index:
        mismatch_letter = get_column_letter(col_index["mismatch_flag"])
        full_range = f"A2:{get_column_letter(len(data_columns))}{max_row}"
        mismatch_fill = PatternFill(
            fill_type="solid",
            start_color="FFF3B0",
            end_color="FFF3B0",
        )
        worksheet.conditional_formatting.add(
            full_range,
            FormulaRule(
                formula=[f"${mismatch_letter}2=1"],
                stopIfTrue=True,
                fill=mismatch_fill,
            ),
        )


def _build_qa_summary(results: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    corr_rows: list[dict[str, float | str]] = []
    for raw_column in ["pys_kiiht_raw", "siv_kiiht_raw", "nyo_kiiht_raw"]:
        corr_rows.append(
            {
                "metric": raw_column,
                "spearman_vs_anomaly_score": float(
                    results[raw_column].corr(results["anomaly_score"], method="spearman")
                ),
                "pearson_vs_anomaly_score": float(
                    results[raw_column].corr(results["anomaly_score"], method="pearson")
                ),
            }
        )
    correlations = pd.DataFrame(corr_rows)

    category_means = (
        results.groupby("anomaly_category", observed=False)[
            ["pys_kiiht_raw", "siv_kiiht_raw", "nyo_kiiht_raw", "anomaly_score"]
        ]
        .mean()
        .reindex(QUALITY_LABELS)
        .reset_index()
    )
    category_counts = (
        results["anomaly_category"]
        .value_counts()
        .reindex(QUALITY_LABELS, fill_value=0)
        .rename_axis("anomaly_category")
        .reset_index(name="count")
    )
    category_summary = category_counts.merge(category_means, on="anomaly_category", how="left")

    top_worst = results.nlargest(20, "anomaly_score")[
        [
            "pys_kiiht_raw",
            "siv_kiiht_raw",
            "nyo_kiiht_raw",
            "anomaly_score",
            "anomaly_category",
            "mismatch_flag",
        ]
    ].copy()

    return correlations, category_summary, top_worst


def _write_excel_results(results: pd.DataFrame, results_path: Path) -> None:
    data_export = results.copy()
    data_export["anomaly_category"] = data_export["anomaly_category"].astype(str)

    correlations, category_summary, top_worst = _build_qa_summary(data_export)

    results_path.parent.mkdir(exist_ok=True)
    with pd.ExcelWriter(results_path, engine="openpyxl") as writer:
        data_export.to_excel(writer, sheet_name="Data", index=False)

        start_row = 0
        correlations.to_excel(writer, sheet_name="QA_summary", index=False, startrow=start_row)
        start_row += len(correlations) + 3
        category_summary.to_excel(writer, sheet_name="QA_summary", index=False, startrow=start_row)
        start_row += len(category_summary) + 3
        top_worst.to_excel(writer, sheet_name="QA_summary", index=False, startrow=start_row)

        ws_data = writer.sheets["Data"]
        ws_qa = writer.sheets["QA_summary"]
        max_row = len(data_export) + 1
        _apply_data_sheet_formatting(ws_data, list(data_export.columns), max_row)
        ws_qa.freeze_panes = "A2"
        for cell in ws_qa[1]:
            cell.font = Font(bold=True)


# -------------------------
# MODEL TESTING / INFERENCE
# -------------------------
def step_07_model_testing(
    model: object,
    all_road_scaled: pd.DataFrame,
    all_road_unscaled: pd.DataFrame,
) -> pd.DataFrame:
    """
    Tests the trained model on all-road data and writes result Excel.

    Args:
        model: Fitted model with predict() and decision_function()
        all_road_scaled: Scaled all road data
        all_road_unscaled: Unscaled all road data (same features/order as all_road_scaled)

    Returns:
        DataFrame with inference outputs and anomaly scores

    Raises:
        TypeError: If inputs are invalid types
        ValueError: If inputs are empty or mismatched
    """
    if not hasattr(model, "predict") or not hasattr(model, "decision_function"):
        raise TypeError(
            "model must provide predict() and decision_function() methods"
        )
    if not isinstance(all_road_scaled, pd.DataFrame):
        raise TypeError("all_road_scaled must be a pandas DataFrame")
    if not isinstance(all_road_unscaled, pd.DataFrame):
        raise TypeError("all_road_unscaled must be a pandas DataFrame")
    if all_road_scaled.empty or all_road_unscaled.empty:
        raise ValueError("all_road_scaled and all_road_unscaled cannot be empty")
    if len(all_road_scaled) != len(all_road_unscaled):
        raise ValueError(
            "all_road_scaled and all_road_unscaled row counts must match"
        )

    print()
    print("------------------------------------------------------------")
    print("Model testing started. Running inference on all-road data...")
    print()

    features = all_road_scaled.columns.tolist()
    scores = model.decision_function(all_road_scaled[features])
    anomaly_scores = -scores
    predictions = model.predict(all_road_scaled[features]).astype(int)

    results = all_road_scaled.copy()
    results["pys_kiiht_raw"] = all_road_unscaled["pys_kiiht"].values
    results["siv_kiiht_raw"] = all_road_unscaled["siv_kiiht"].values
    results["nyo_kiiht_raw"] = all_road_unscaled["nyo_kiiht"].values
    results["anomaly_prediction"] = predictions
    # anomaly_score semantics: larger value = more anomalous / worse road
    results["anomaly_score"] = anomaly_scores
    anomaly_score_shifted = anomaly_scores - float(np.min(anomaly_scores))
    anomaly_score_log = np.log1p(anomaly_score_shifted)
    results["anomaly_score_log"] = anomaly_score_log
    results["anomaly_type"] = results["anomaly_prediction"].map(
        {1: "Normal", -1: "Anomaly"}
    )
    # Robust rank-based categorization avoids degenerate-bin failures when
    # quantile boundaries are equal in discrete/quantized score distributions.
    score_rank = results["anomaly_score_log"].rank(method="average", pct=True)
    conditions = [
        score_rank <= QUALITY_BINS[0],
        score_rank <= QUALITY_BINS[1],
        score_rank <= QUALITY_BINS[2],
        score_rank <= QUALITY_BINS[3],
    ]
    choices = QUALITY_LABELS[:-1]
    category_values = np.select(conditions, choices, default=QUALITY_LABELS[-1])
    priority_category = pd.Categorical(
        category_values,
        categories=QUALITY_LABELS,
        ordered=True,
    )
    results["anomaly_category"] = priority_category

    # Flag contradictory extremes for manual QA:
    # high acceleration but low anomaly score, or low acceleration but high anomaly score.
    accel_sum = (
        results["pys_kiiht_raw"]
        + results["siv_kiiht_raw"]
        + results["nyo_kiiht_raw"]
    )
    accel_rank = accel_sum.rank(method="average", pct=True)
    score_rank_for_mismatch = results["anomaly_score"].rank(method="average", pct=True)
    high_accel_low_score = (accel_rank >= (1.0 - MISMATCH_Q)) & (
        score_rank_for_mismatch <= MISMATCH_Q
    )
    low_accel_high_score = (accel_rank <= MISMATCH_Q) & (
        score_rank_for_mismatch >= (1.0 - MISMATCH_Q)
    )
    results["mismatch_flag"] = (high_accel_low_score | low_accel_high_score).astype(int)

    # Keep quantile values for reporting visibility (not used for binning).
    q05 = float(results["anomaly_score_log"].quantile(QUALITY_BINS[0]))
    q20 = float(results["anomaly_score_log"].quantile(QUALITY_BINS[1]))
    q50 = float(results["anomaly_score_log"].quantile(QUALITY_BINS[2]))
    q80 = float(results["anomaly_score_log"].quantile(QUALITY_BINS[3]))

    anomaly_count = int((predictions == -1).sum())
    normal_count = int((predictions == 1).sum())
    total_count = len(predictions)

    print("Inference summary (all-road data):")
    print("  Dataset used: all_road_scaled")
    print(f"  Inference rows (all_road_scaled): {total_count}")
    print(f"  Normal roads: {normal_count} ({normal_count / total_count * 100:.1f}%)")
    print(
        f"  Anomalous roads: {anomaly_count} ({anomaly_count / total_count * 100:.1f}%)"
    )
    print(f"  Actual contamination rate: {anomaly_count / total_count:.4f}")
    print()

    print("Anomaly score distribution (all-road data, larger = worse):")
    score_series = pd.Series(anomaly_scores)
    print(f"  Min: {float(score_series.min()):.6f}")
    print(f"  Q01: {float(score_series.quantile(0.01)):.6f}")
    print(f"  Q05: {float(score_series.quantile(0.05)):.6f}")
    print(f"  Median: {float(score_series.median()):.6f}")
    print(f"  Q95: {float(score_series.quantile(0.95)):.6f}")
    print(f"  Max: {float(score_series.max()):.6f}")
    print()
    log_score_series = pd.Series(anomaly_score_log)
    print("Anomaly score (log1p shifted) distribution:")
    print(f"  Min: {float(log_score_series.min()):.6f}")
    print(f"  Q01: {float(log_score_series.quantile(0.01)):.6f}")
    print(f"  Q05: {float(log_score_series.quantile(0.05)):.6f}")
    print(f"  Median: {float(log_score_series.median()):.6f}")
    print(f"  Q95: {float(log_score_series.quantile(0.95)):.6f}")
    print(f"  Max: {float(log_score_series.max()):.6f}")
    print()
    print("Road quality category thresholds (from anomaly_score_log quantiles):")
    print(f"  <= Q05 ({q05:.6f}): Excellent")
    print(f"  <= Q20 ({q20:.6f}): Good")
    print(f"  <= Q50 ({q50:.6f}): Fair")
    print(f"  <= Q80 ({q80:.6f}): Poor")
    print(f"  >  Q80 ({q80:.6f}): Critical")
    print()
    print("Priority category distribution:")
    category_counts = (
        results["anomaly_category"]
        .value_counts()
        .reindex(QUALITY_LABELS, fill_value=0)
    )
    for category in QUALITY_LABELS:
        count = int(category_counts[category])
        share = (count / total_count) * 100
        print(f"  {category}: {count} ({share:.1f}%)")
    priority_count = int(results["anomaly_category"].isin(PRIORITY_LABELS).sum())
    print(
        "  is_priority (Poor/Critical): "
        f"{priority_count} ({priority_count / total_count * 100:.1f}%)"
    )
    mismatch_count = int(results["mismatch_flag"].sum())
    print(f"  mismatch_flag rows: {mismatch_count} ({mismatch_count / total_count * 100:.3f}%)")
    print()

    print("------------------------------------------------------------")
    print("Building Excel results table next, please wait...")
    results_path = OUTPUT_DIR / "anomaly_results.xlsx"
    results = results.sort_values("anomaly_score", ascending=False, kind="mergesort").reset_index(drop=True)
    _write_excel_results(results, results_path)
    print(f"Results saved to: {results_path}")

    return results
